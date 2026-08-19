import pandas as pd
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from typing import Dict, List, Any
from core.db.models import ParserProduct, ParserCharacteristics, ParserGroupCharacteristics

def export_products_to_excel(db: Session, filename: str = "products_export.xlsx") -> None:
    """
    Экспортирует товары с характеристиками в Excel

    Args:
        db: Сессия БД
        filename: Имя файла для сохранения
    """

    # 1. Получаем все товары
    products = db.query(ParserProduct).all()

    if not products:
        logging.warning("Нет товаров для экспорта")
        return

    # 2. Собираем данные
    data = []
    all_characteristics = set()  # Множество всех уникальных характеристик

    # Сначала проходим по всем товарам, чтобы собрать все возможные характеристики
    for product in products:
        for char in product.characteristics:
            if char.group:
                all_characteristics.add(char.group.name)

    # Сортируем характеристики для удобства
    all_characteristics = sorted(list(all_characteristics))

    # Формируем данные для каждого товара
    for product in products:
        # Базовые поля товара
        row = {
            "ID": product.id,
            "ID_MS": product.id_ms,
            "Код_МС": product.code_ms,
            "URL": product.url,
            "Изображения": product.images,
            "Описание": product.description[:500] + "..." if len(product.description) > 500 else product.description,
            "Бренд": product.brand,
            "Цена": product.prices,
            "Дата_обновления": product.updated_at.strftime("%Y-%m-%d %H:%M:%S") if product.updated_at else ""
        }

        # Заполняем характеристики
        # Создаем словарь характеристик для текущего товара
        chars_dict = {}
        for char in product.characteristics:
            if char.group:
                chars_dict[char.group.name] = char.value

        # Добавляем все характеристики в строку
        for char_name in all_characteristics:
            row[char_name] = chars_dict.get(char_name, "")

        data.append(row)

    # 3. Создаем DataFrame
    df = pd.DataFrame(data)

    # 4. Сохраняем в Excel с форматированием
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Товары', index=False)

        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Товары']

        # Форматируем заголовки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Автоматически подгоняем ширину колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Ограничиваем ширину 50 символов
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Замораживаем заголовок
        worksheet.freeze_panes = 'A2'

    logging.info(f"✅ Экспортировано {len(data)} товаров в файл {filename}")
    print(f"\n✅ Экспорт завершен!")
    print(f"📊 Всего товаров: {len(data)}")
    print(f"📋 Характеристик: {len(all_characteristics)}")
    print(f"💾 Файл сохранен: {filename}")


def export_products_to_excel_with_characteristics(
        db: Session,
        filename: str = "products_export.xlsx",
        domain: str = None
) -> None:
    """
    Экспортирует товары с характеристиками в Excel с возможностью фильтрации по домену

    Args:
        db: Сессия БД
        filename: Имя файла для сохранения
        domain: Фильтр по домену (например, "vseinstrumenti.ru")
    """

    # Базовый запрос
    query = db.query(ParserProduct)

    # Фильтруем по домену если указан
    if domain:
        query = query.join(ParserProduct.characteristics).join(
            ParserCharacteristics.group
        ).filter(ParserGroupCharacteristics.domen == domain).distinct()

    products = query.all()

    if not products:
        logging.warning("Нет товаров для экспорта")
        return

    # Собираем все характеристики
    all_characteristics = set()
    for product in products:
        for char in product.characteristics:
            if char.group:
                all_characteristics.add(char.group.name)

    all_characteristics = sorted(list(all_characteristics))

    # Формируем данные
    data = []
    for product in products:
        row = {
            "ID": product.id,
            "ID_MS": product.id_ms,
            "Код_МС": product.code_ms,
            "URL": product.url,
            "Бренд": product.brand,
            "Цена": product.prices,
            "Дата_обновления": product.updated_at.strftime("%Y-%m-%d %H:%M:%S") if product.updated_at else ""
        }

        # Добавляем описание (обрезаем для читаемости)
        desc = product.description or ""
        row["Описание"] = desc[:300] + "..." if len(desc) > 300 else desc

        # Добавляем изображения (первые 3)
        try:
            import json
            images = json.loads(product.images) if product.images else []
            row["Изображений"] = len(images)
            row["Первое_изображение"] = images[0].get("url", "")[:100] if images else ""
        except:
            row["Изображений"] = 0
            row["Первое_изображение"] = ""

        # Характеристики
        chars_dict = {}
        for char in product.characteristics:
            if char.group:
                chars_dict[char.group.name] = char.value

        for char_name in all_characteristics:
            row[char_name] = chars_dict.get(char_name, "")

        data.append(row)

    # Создаем DataFrame
    df = pd.DataFrame(data)

    # Сохраняем в Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Товары', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Товары']

        # Форматирование
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        worksheet.freeze_panes = 'A2'

    print(f"\n✅ Экспорт завершен!")
    print(f"📊 Всего товаров: {len(data)}")
    print(f"📋 Всего характеристик: {len(all_characteristics)}")
    print(f"💾 Файл: {filename}")


def export_products_with_characteristics_as_columns(db: Session, filename: str = "products_export.xlsx"):
    """
    Альтернативный вариант: выгружает с характеристиками в виде отдельных колонок
    и сохраняет в несколько листов
    """

    products = db.query(ParserProduct).all()

    if not products:
        print("Нет товаров для экспорта")
        return

    # Собираем все характеристики
    all_chars = set()
    for product in products:
        for char in product.characteristics:
            if char.group:
                all_chars.add(char.group.name)

    all_chars = sorted(list(all_chars))

    # Данные для основного листа
    main_data = []
    for product in products:
        row = {
            "ID": product.id,
            "Код_МС": product.code_ms,
            "Название": f"{product.brand} {product.code_ms}" if product.brand else product.code_ms,
            "Бренд": product.brand,
            "Цена": product.prices,
            "Описание": (product.description or "")[:200]
        }
        main_data.append(row)

    # Данные для листа характеристик
    chars_data = []
    for product in products:
        for char in product.characteristics:
            if char.group:
                chars_data.append({
                    "Код_МС": product.code_ms,
                    "Характеристика": char.group.name,
                    "Значение": char.value,
                    "Домен": char.group.domen
                })

    # Создаем Excel с несколькими листами
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Лист с товарами
        pd.DataFrame(main_data).to_excel(writer, sheet_name='Товары', index=False)

        # Лист с характеристиками (список)
        pd.DataFrame(chars_data).to_excel(writer, sheet_name='Характеристики', index=False)

        # Лист со сводной таблицей (товар-характеристика)
        pivot_data = []
        for product in products:
            row = {"Код_МС": product.code_ms}
            for char in product.characteristics:
                if char.group:
                    row[char.group.name] = char.value
            pivot_data.append(row)

        # Заполняем пустые значения
        df_pivot = pd.DataFrame(pivot_data)
        df_pivot = df_pivot.fillna("")
        df_pivot.to_excel(writer, sheet_name='Сводная', index=False)

    print(f"✅ Экспортировано {len(products)} товаров")
    print(f"📊 Характеристик: {len(all_chars)}")
    print(f"💾 Сохранено в {filename}")


# Пример использования
if __name__ == "__main__":
    from core.db.connection import get_db_session

    with get_db_session() as db:
        # Вариант 1: базовая выгрузка
        export_products_to_excel(db, "products_full.xlsx")

        # Вариант 2: с фильтром по домену
        # export_products_to_excel_with_characteristics(
        #     db,
        #     "products_vseinstrumenti.xlsx",
        #     domain="vseinstrumenti.ru"
        # )

        # Вариант 3: с несколькими листами
        # export_products_with_characteristics_as_columns(db, "products_multi.xlsx")