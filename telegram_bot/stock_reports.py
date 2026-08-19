from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Callable, Iterable, Sequence
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from app.db import SessionLocal
from app.models import (
    OzonProduct,
    OzonStockSnapshot,
    OzonWarehouse,
    OzonWarehouseStockSnapshot,
    WBFboStockSnapshot,
    WBFboWarehouse,
    WBFBSStockSnapshot,
    WBFBSWarehouse,
    WBProduct,
    WBProductSize,
)


class StockSnapshotNotFound(RuntimeError):
    pass


def _excel_value(value: Any) -> Any:
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(ZoneInfo("Europe/Moscow")).replace(tzinfo=None)
    return value


def build_workbook(sheets: Sequence[tuple[str, Sequence[str], Iterable[Sequence[Any]]]]) -> bytes:
    """Create an xlsx payload in memory; no temporary file is used."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for title, headers, rows in sheets:
        sheet = workbook.create_sheet(title[:31])
        sheet.append(list(headers))
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        widths = [len(str(header)) for header in headers]
        for row in rows:
            values = [_excel_value(value) for value in row]
            sheet.append(values)
            for index, value in enumerate(values):
                if value is not None:
                    widths[index] = min(60, max(widths[index], len(str(value))))
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = min(60, width + 2)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


class StockExcelReportService:
    def __init__(self, *, session_factory: Callable[..., Any] = SessionLocal) -> None:
        self.session_factory = session_factory

    def wb(self, snapshot_date: date) -> tuple[str, bytes, str]:
        with self.session_factory() as session:
            fbs = session.query(
                WBFBSStockSnapshot.snapshot_date,
                WBFBSStockSnapshot.captured_at,
                WBProduct.nm_id,
                WBProduct.vendor_code,
                WBProduct.title,
                WBProduct.brand,
                WBProductSize.tech_size,
                WBProductSize.wb_size,
                WBFBSStockSnapshot.sku,
                WBFBSWarehouse.name,
                WBFBSStockSnapshot.quantity,
            ).join(WBProductSize, WBProductSize.id == WBFBSStockSnapshot.size_id).join(
                WBProduct, WBProduct.id == WBProductSize.product_id
            ).join(WBFBSWarehouse, WBFBSWarehouse.id == WBFBSStockSnapshot.warehouse_id).filter(
                WBFBSStockSnapshot.snapshot_date == snapshot_date,
                WBFBSStockSnapshot.quantity > 0,
            ).order_by(WBProduct.vendor_code, WBFBSWarehouse.name).all()
            fbo = session.query(
                WBFboStockSnapshot.snapshot_date,
                WBFboStockSnapshot.captured_at,
                WBProduct.nm_id,
                WBProduct.vendor_code,
                WBProduct.title,
                WBProduct.brand,
                WBProductSize.tech_size,
                WBProductSize.wb_size,
                WBFboWarehouse.name,
                WBFboWarehouse.region_name,
                WBFboStockSnapshot.quantity,
                WBFboStockSnapshot.in_way_to_client,
                WBFboStockSnapshot.in_way_from_client,
            ).join(WBProductSize, WBProductSize.id == WBFboStockSnapshot.size_id).join(
                WBProduct, WBProduct.id == WBProductSize.product_id
            ).join(WBFboWarehouse, WBFboWarehouse.id == WBFboStockSnapshot.warehouse_id).filter(
                WBFboStockSnapshot.snapshot_date == snapshot_date,
                WBFboStockSnapshot.quantity > 0,
            ).order_by(WBProduct.vendor_code, WBFboWarehouse.name).all()
        if not fbs and not fbo:
            raise StockSnapshotNotFound(f"WB stock snapshot for {snapshot_date.isoformat()} was not found")
        content = build_workbook([
            ("FBS", ("Дата", "Снято", "nmID", "Артикул", "Товар", "Бренд", "Размер", "Размер WB", "Баркод", "Склад", "Остаток"), fbs),
            ("FBO", ("Дата", "Снято", "nmID", "Артикул", "Товар", "Бренд", "Размер", "Размер WB", "Склад", "Регион", "Остаток", "К клиенту", "От клиента"), fbo),
        ])
        filename = f"wb_stocks_{snapshot_date.isoformat()}.xlsx"
        return filename, content, f"Остатки Wildberries на {snapshot_date:%d.%m.%Y} (00:00 МСК)"

    def ozon(self, snapshot_date: date) -> tuple[str, bytes, str]:
        with self.session_factory() as session:
            warehouses = session.query(
                OzonWarehouseStockSnapshot.snapshot_date,
                OzonWarehouseStockSnapshot.captured_at,
                OzonWarehouseStockSnapshot.product_id,
                OzonWarehouseStockSnapshot.offer_id,
                OzonWarehouseStockSnapshot.sku,
                OzonProduct.name,
                OzonWarehouseStockSnapshot.stock_type,
                OzonWarehouse.ozon_warehouse_id,
                func.coalesce(OzonWarehouse.name, "Название не получено"),
                func.coalesce(OzonWarehouse.cluster_name, "Кластер не получен"),
                OzonWarehouseStockSnapshot.present,
                OzonWarehouseStockSnapshot.reserved,
            ).join(OzonWarehouse, OzonWarehouse.id == OzonWarehouseStockSnapshot.warehouse_id).outerjoin(
                OzonProduct, OzonProduct.product_id == OzonWarehouseStockSnapshot.product_id
            ).filter(
                OzonWarehouseStockSnapshot.snapshot_date == snapshot_date,
                OzonWarehouseStockSnapshot.present > 0,
            ).order_by(
                OzonWarehouseStockSnapshot.offer_id, OzonWarehouseStockSnapshot.stock_type, OzonWarehouse.name
            ).all()
            aggregate = session.query(
                OzonStockSnapshot.snapshot_date,
                OzonStockSnapshot.captured_at,
                OzonStockSnapshot.product_id,
                OzonStockSnapshot.offer_id,
                OzonProduct.name,
                OzonStockSnapshot.stock_type,
                OzonStockSnapshot.present,
                OzonStockSnapshot.reserved,
            ).outerjoin(OzonProduct, OzonProduct.product_id == OzonStockSnapshot.product_id).filter(
                OzonStockSnapshot.snapshot_date == snapshot_date,
                OzonStockSnapshot.present > 0,
            ).order_by(OzonStockSnapshot.offer_id, OzonStockSnapshot.stock_type).all()
        if not warehouses and not aggregate:
            raise StockSnapshotNotFound(f"Ozon stock snapshot for {snapshot_date.isoformat()} was not found")
        content = build_workbook([
            ("По складам", ("Дата", "Снято", "Product ID", "Offer ID", "SKU", "Товар", "Тип", "ID склада Ozon", "Название склада", "Кластер", "В наличии", "В резерве"), warehouses),
            ("Агрегат", ("Дата", "Снято", "Product ID", "Offer ID", "Товар", "Тип", "В наличии", "В резерве"), aggregate),
        ])
        filename = f"ozon_stocks_{snapshot_date.isoformat()}.xlsx"
        return filename, content, f"Остатки Ozon на {snapshot_date:%d.%m.%Y} (00:00 МСК)"
