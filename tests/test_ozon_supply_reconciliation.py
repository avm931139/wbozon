from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db import Base
from app.models import (
    OzonFBOSupplyAct,
    OzonFBOSupplyActItem,
    OzonFBOSupplyDeclaredItem,
)
from ozon.exceptions import OzonHTTPError, OzonRateLimitError
from ozon.services.supply_reconciliation_service import OzonSupplyReconciliationService
from ozon.services.sync_service import OzonSyncService
from ozon.supply_reconciliation_report import OzonSupplyReconciliationExcelReport
from ozon.supplies import OzonSuppliesAPI


class QueueClient:
    def __init__(self, payloads):
        self.payloads = iter(payloads)
        self.calls = []

    def post(self, path, *, json_body=None, retries=3):
        self.calls.append((path, json_body, retries))
        return next(self.payloads)


def test_supply_bundle_paginates_and_act_methods_use_current_contracts():
    client = QueueClient([
        {"items": [{"sku": 1}], "has_next": True, "last_id": "1"},
        {"items": [{"sku": 2}], "has_next": False, "last_id": "2"},
        {"supplies_acts": []},
        {"supply_id": 501, "supply_acts": [], "skus_defects": []},
    ])
    api = OzonSuppliesAPI(client)

    assert [item["sku"] for item in api.bundle("bundle-1", limit=1)] == [1, 2]
    assert api.act_summary(1001) == {"supplies_acts": []}
    assert api.act_products(501)["supply_id"] == 501
    assert client.calls == [
        (
            "/v1/supply-order/bundle",
            {
                "bundle_ids": ["bundle-1"],
                "limit": 1,
                "last_id": "",
                "is_asc": True,
                "sort_field": "SKU",
            },
            6,
        ),
        (
            "/v1/supply-order/bundle",
            {
                "bundle_ids": ["bundle-1"],
                "limit": 1,
                "last_id": "1",
                "is_asc": True,
                "sort_field": "SKU",
            },
            6,
        ),
        ("/v1/supply-order/act/summary/get", {"order_id": 1001}, 6),
        ("/v1/supply-order/act/product/get", {"supply_id": 501}, 6),
    ]


class FakeSupplyAPI:
    def __init__(self):
        self.bundle_calls = []
        self.product_calls = []

    def list(self):
        return [{"supply_order_id": 900}, {"supply_order_id": 1001}]

    def get(self, order_id):
        if order_id == 900:
            return {"order_id": 900, "created_date": "2025-12-31T10:00:00Z", "supplies": []}
        return {
            "order_id": 1001,
            "created_date": "2026-01-10T10:00:00Z",
            "state": "IN_TRANSIT",
            "supplies": [
                {
                    "supply_id": 501,
                    "bundle_id": "bundle-1",
                    "state": "COMPLETED",
                    "storage_warehouse": {"warehouse_id": 77, "name": "Хоругвино"},
                },
                {
                    "supply_id": 502,
                    "bundle_id": "bundle-draft",
                    "state": "READY_TO_SUPPLY",
                },
            ],
        }

    def bundle(self, bundle_id):
        self.bundle_calls.append(bundle_id)
        return [{
            "sku": 100,
            "product_id": 200,
            "offer_id": "offer-100",
            "name": "Товар",
            "barcode": "4600000000000",
            "quantity": 10,
            "quant": 1,
            "shipment_type": "DIRECT",
            "placement_zone": "SORT",
            "tags": [],
        }]

    def act_summary(self, order_id):
        assert order_id == 1001
        return {"supplies_acts": [{
            "supply_id": 501,
            "is_agreement_completed": True,
            "supply_acts": [
                {
                    "act_id": 701,
                    "act_number": "ACT-701",
                    "act_state": "ACCEPTED",
                    "created_date": "2026-01-15",
                    "type": "ACCEPTANCE",
                    "summary": {
                        "declared_quantity": 10,
                        "fact_quantity": 8,
                        "approved_quantity": 8,
                        "sku_quantity": 1,
                        "unidentified_quantity": 0,
                    },
                },
                {
                    "act_id": 702,
                    "act_number": "ACT-702",
                    "act_state": "ACCEPTED",
                    "created_date": "2026-01-15",
                    "type": "SURPLUS",
                    "summary": {
                        "declared_quantity": 0,
                        "fact_quantity": 2,
                        "approved_quantity": 2,
                        "sku_quantity": 1,
                        "unidentified_quantity": 0,
                    },
                },
            ],
        }]}

    def act_products(self, supply_id):
        self.product_calls.append(supply_id)
        return {
            "supply_id": supply_id,
            "skus_defects": [],
            "supply_acts": [
                {
                    "act_id": 701,
                    "type": "ACCEPTANCE",
                    "unidentified_quantity": 0,
                    "items": [{
                        "declared_quantity": 10,
                        "fact_quantity": 8,
                        "approved_quantity": 8,
                        "sku_info": {"sku": 100, "offer_id": "offer-100", "name": "Товар"},
                    }],
                },
                {
                    "act_id": 702,
                    "type": "SURPLUS",
                    "unidentified_quantity": 0,
                    "items": [{
                        "declared_quantity": 0,
                        "fact_quantity": 2,
                        "approved_quantity": 2,
                        "sku_info": {"sku": 300, "offer_id": "offer-300", "name": "Излишек"},
                    }],
                },
            ],
        }


def test_reconciliation_persists_sent_and_accepted_quantities_idempotently():
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, future=True)
    api = FakeSupplyAPI()
    service = OzonSupplyReconciliationService(
        api=api,
        session_factory=session_factory,
        history_from=date(2026, 1, 1),
        request_pause_seconds=0,
    )

    result = service.sync_all()
    assert result["orders_scanned"] == 2
    assert result["orders_in_period"] == 1
    assert result["supplies_sent"] == 1
    assert result["declared_items"] == 1
    assert result["acts"] == 2
    assert result["act_items"] == 2
    assert result["failed"] == 0
    assert api.bundle_calls == ["bundle-1"]
    assert api.product_calls == [501]

    with session_factory() as session:
        declared = session.query(OzonFBOSupplyDeclaredItem).one()
        assert declared.declared_quantity == 10
        assert declared.storage_warehouse_name == "Хоругвино"
        acceptance = session.query(OzonFBOSupplyActItem).filter_by(
            act_type="ACCEPTANCE",
            sku=100,
        ).one()
        assert acceptance.fact_quantity == 8
        surplus = session.query(OzonFBOSupplyActItem).filter_by(
            act_type="SURPLUS",
            sku=300,
        ).one()
        assert surplus.fact_quantity == 2
        assert session.query(OzonFBOSupplyAct).count() == 2

    service.sync_all()
    with session_factory() as session:
        assert session.query(OzonFBOSupplyDeclaredItem).count() == 1
        assert session.query(OzonFBOSupplyAct).count() == 2
        assert session.query(OzonFBOSupplyActItem).count() == 2


def test_reconciliation_is_an_independent_ozon_task():
    assert "supply_reconciliation" in OzonSyncService.task_names()


def test_missing_acceptance_act_is_normal_for_supply_in_transit():
    class InTransitAPI(FakeSupplyAPI):
        def act_summary(self, order_id):
            raise OzonHTTPError("Ozon API returned HTTP 404: supply act not found")

        def act_products(self, supply_id):
            raise AssertionError("act products must not be requested without a summary")

    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, future=True)
    result = OzonSupplyReconciliationService(
        api=InTransitAPI(),
        session_factory=session_factory,
        history_from=date(2026, 1, 1),
        request_pause_seconds=0,
    ).sync_all()

    assert result["declared_items"] == 1
    assert result["acts_unavailable"] == 1
    assert result["failed"] == 0
    assert "reconciliation_error" not in result


def test_rate_limit_retries_the_same_act_without_losing_the_run():
    class RateLimitedOnceAPI(FakeSupplyAPI):
        def __init__(self):
            super().__init__()
            self.summary_calls = 0

        def act_summary(self, order_id):
            self.summary_calls += 1
            if self.summary_calls == 1:
                raise OzonRateLimitError("Ozon API rate limit exceeded")
            return super().act_summary(order_id)

    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, future=True)
    api = RateLimitedOnceAPI()
    sleeps = []
    result = OzonSupplyReconciliationService(
        api=api,
        session_factory=session_factory,
        history_from=date(2026, 1, 1),
        request_pause_seconds=0,
        rate_limit_retries=2,
        rate_limit_backoff_seconds=7,
        sleeper=sleeps.append,
    ).sync_all()

    assert result["failed"] == 0
    assert result["act_items"] == 2
    assert api.summary_calls == 2
    assert sleeps == [7]


def test_exhausted_rate_limit_keeps_declared_quantities_and_reports_partial():
    class AlwaysRateLimitedAPI(FakeSupplyAPI):
        def __init__(self):
            super().__init__()
            self.summary_calls = 0

        def act_summary(self, order_id):
            self.summary_calls += 1
            raise OzonRateLimitError("Ozon API rate limit exceeded")

    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, future=True)
    api = AlwaysRateLimitedAPI()
    sleeps = []
    result = OzonSupplyReconciliationService(
        api=api,
        session_factory=session_factory,
        history_from=date(2026, 1, 1),
        request_pause_seconds=0,
        rate_limit_retries=2,
        rate_limit_backoff_seconds=3,
        sleeper=sleeps.append,
    ).sync_all()

    assert result["declared_items"] == 1
    assert result["act_items"] == 0
    assert result["failed"] == 1
    assert "act-summary:1001" in result["reconciliation_error"]
    assert api.summary_calls == 3
    assert sleeps == [3, 6]


def test_excel_report_contains_product_and_supply_totals(tmp_path):
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, future=True)
    OzonSupplyReconciliationService(
        api=FakeSupplyAPI(),
        session_factory=session_factory,
        history_from=date(2026, 1, 1),
        request_pause_seconds=0,
    ).sync_all()

    report = OzonSupplyReconciliationExcelReport(session_factory=session_factory)
    content, product_rows, supply_rows = report.build()
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)

    assert workbook.sheetnames == ["По товарам", "По поставкам"]
    assert product_rows == 2
    assert supply_rows == 1
    products = list(workbook["По товарам"].iter_rows(values_only=True))
    headers = {value: index for index, value in enumerate(products[0])}
    accepted = next(row for row in products[1:] if row[headers["SKU"]] == 100)
    assert accepted[headers["Номер заявки"]] == 1001
    assert accepted[headers["ID поставки"]] == 501
    assert accepted[headers["Отправлено"]] == 10
    assert accepted[headers["Принято фактически"]] == 8
    assert accepted[headers["Расхождение принято−отправлено"]] == -2
    supplies = list(workbook["По поставкам"].iter_rows(values_only=True))
    assert supplies[1][supplies[0].index("Отправлено")] == 10
    assert supplies[1][supplies[0].index("Принято фактически")] == 8
    workbook.close()

    result = report.save(tmp_path / "report.xlsx")
    assert result["product_rows"] == 2
    assert result["supply_rows"] == 1
    assert (tmp_path / "report.xlsx").read_bytes().startswith(b"PK")
