from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
import pytest
import requests
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from zoneinfo import ZoneInfo

from app.db import Base
from app.models import YandexMarketStockSnapshot
from telegram_bot.client import TelegramClient, TelegramError, split_text
from telegram_bot.__main__ import send_stock_files
from telegram_bot.reports import TelegramReportService
from telegram_bot.scheduler import TelegramReportScheduler
from telegram_bot.stock_reports import (
    StockExcelReportService,
    StockSnapshotNotFound,
    build_workbook,
)


class FakeResponse:
    def __init__(self, message_id): self.message_id = message_id
    def raise_for_status(self): return None
    def json(self): return {"ok": True, "result": {"message_id": self.message_id}}


class FakeHTTPSession:
    def __init__(self): self.calls = []; self.proxies = {}
    def post(self, url, **kwargs):
        self.calls.append((url, kwargs)); return FakeResponse(len(self.calls))


def sales_data():
    return {
        "orders_placed": 14, "orders_amount": "70010.25", "orders_from_period_now_cancelled": 2,
        "cancellations_registered": 6, "buyouts": 14, "buyouts_amount": "40227", "returns": 1,
        "returns_amount": "1000", "net_buyouts": 13, "net_buyouts_amount": "39227",
        "unknown_operations": 0, "operations_without_order_row": 1,
        "fulfillment": {"orders": {"fbs": 5, "fbo": 9}, "buyouts": {"fbs": 4, "fbo": 10}},
        "accounting_covers_period": False, "accounting_report_through": "2026-08-02",
    }


def test_split_text_and_client_send_every_chunk():
    assert all(len(chunk) <= 10 for chunk in split_text("first\n" + "x" * 30, limit=10))
    session = FakeHTTPSession(); client = TelegramClient("secret", "-1001", session=session)
    assert client.send_text("a" * 4000) == [1, 2]
    assert all(call[1]["json"]["chat_id"] == "-1001" for call in session.calls)
    assert all("secret" not in str(call[1]) for call in session.calls)


def test_client_sends_document_as_multipart_without_disk_file():
    session = FakeHTTPSession(); client = TelegramClient("secret", "-1001", session=session)
    assert client.send_document("stocks.xlsx", b"xlsx", caption="Stocks") == 1
    url, kwargs = session.calls[0]
    assert url.endswith("/sendDocument")
    assert kwargs["data"] == {"chat_id": "-1001", "caption": "Stocks"}
    assert kwargs["files"]["document"][0:2] == ("stocks.xlsx", b"xlsx")


def test_client_applies_proxy_only_to_its_session():
    session = FakeHTTPSession()
    proxy_url = "socks5h://127.0.0.1:1080"
    TelegramClient("secret", "-1001", proxy_url=proxy_url, session=session)
    assert session.proxies == {"http": proxy_url, "https": proxy_url}


def test_client_rejects_invalid_proxy_url():
    with pytest.raises(ValueError, match="WB_TG_PROXY_URL"):
        TelegramClient("secret", "-1001", proxy_url="ftp://proxy.example", session=FakeHTTPSession())


def test_transport_error_redacts_bot_token():
    token = "123456:very-secret-token"

    class FailingSession(FakeHTTPSession):
        def post(self, url, **kwargs):
            raise requests.ConnectionError(f"connection failed for {url}")

    client = TelegramClient(token, "-1001", session=FailingSession())
    with pytest.raises(TelegramError) as error:
        client.send_text("hello")
    assert token not in str(error.value)
    assert "<redacted>" in str(error.value)


def test_build_workbook_returns_in_memory_xlsx():
    payload = build_workbook([("Остатки", ("Дата", "SKU", "Количество"), [(date(2026, 8, 19), "sku-1", 7)])])
    workbook = load_workbook(BytesIO(payload), read_only=True)
    sheet = workbook["Остатки"]
    assert list(sheet.values) == [("Дата", "SKU", "Количество"), (datetime(2026, 8, 19), "sku-1", 7)]
    workbook.close()


class FakeStockDispatcher:
    def __init__(self): self.documents = []; self.warnings = []
    def send_document(self, report_type, report_key, factory, **kwargs):
        document = factory(); self.documents.append((report_type, report_key, document)); return {"status": "sent"}
    def send_text_content(self, report_type, report_key, factory, **kwargs):
        text = factory(); self.warnings.append((report_type, report_key, text)); return {"status": "sent"}


class PartiallyMissingStockReports:
    def wb(self, snapshot_date):
        raise StockSnapshotNotFound("missing WB")
    def ozon(self, snapshot_date):
        return "ozon.xlsx", b"xlsx", "Ozon"
    def yandex_market(self, snapshot_date):
        raise StockSnapshotNotFound("missing Yandex Market")


def test_stock_files_warn_when_daily_snapshot_is_missing_but_send_available_file():
    dispatcher = FakeStockDispatcher()
    results = send_stock_files(
        dispatcher,
        date(2026, 8, 20),
        reports=PartiallyMissingStockReports(),
    )
    assert len(results) == 2
    assert len(dispatcher.documents) == 1
    assert dispatcher.documents[0][0] == "stock_excel_ozon"
    assert dispatcher.warnings[0][1] == "stock_warning:2026-08-20"
    assert "Wildberries" in dispatcher.warnings[0][2]
    assert "Яндекс Маркет" in dispatcher.warnings[0][2]
    assert "20.08.2026" in dispatcher.warnings[0][2]


def test_yandex_market_stock_report_contains_summary_and_warehouse_rows():
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, future=True)
    snapshot_date = date(2026, 8, 30)
    with session_factory() as session:
        session.add_all([
            YandexMarketStockSnapshot(
                snapshot_date=snapshot_date,
                captured_at=datetime(2026, 8, 30, 0, 1, tzinfo=ZoneInfo("UTC")),
                campaign_id=149007825,
                warehouse_id=10,
                offer_id="sku-1",
                stock_type="AVAILABLE",
                count=12,
                source_updated_at=datetime(2026, 8, 29, 23, 59, tzinfo=ZoneInfo("UTC")),
                raw_data={},
            ),
            YandexMarketStockSnapshot(
                snapshot_date=snapshot_date,
                captured_at=datetime(2026, 8, 30, 0, 1, tzinfo=ZoneInfo("UTC")),
                campaign_id=149007825,
                warehouse_id=10,
                offer_id="sku-zero",
                stock_type="AVAILABLE",
                count=0,
                source_updated_at=None,
                raw_data={},
            ),
        ])
        session.commit()

    filename, payload, caption = StockExcelReportService(
        session_factory=session_factory
    ).yandex_market(snapshot_date)

    assert filename == "yandex_market_stocks_2026-08-30.xlsx"
    assert "Яндекс Маркета" in caption
    workbook = load_workbook(BytesIO(payload), read_only=True)
    assert workbook.sheetnames == ["Сводка", "По складам"]
    assert list(workbook["Сводка"].values)[1] == (
        datetime(2026, 8, 30),
        149007825,
        "AVAILABLE",
        1,
        12,
    )
    warehouse_rows = list(workbook["По складам"].values)
    assert len(warehouse_rows) == 2
    assert warehouse_rows[1][4:7] == ("sku-1", "AVAILABLE", 12)
    workbook.close()


def test_sales_block_keeps_event_definitions_separate():
    block = TelegramReportService._sales_block("ПРОШЛЫЙ ДЕНЬ", sales_data())
    for text in ("Заказы: 14", "сейчас отменено: 2", "Отмен зарегистрировано в периоде: 6", "Выкупы: 14", "Возвраты: 1", "оперативный"):
        assert text in block


class FakeDispatcher:
    def __init__(self): self.calls = []
    def send(self, report_type, report_key, **kwargs):
        self.calls.append((report_type, report_key)); return {"status": "sent"}


def test_scheduler_builds_daily_and_interval_keys():
    dispatcher = FakeDispatcher()
    scheduler = TelegramReportScheduler(dispatcher, timezone_name="Europe/Moscow", morning_time="09:00", operational_interval_seconds=10800)
    scheduler.run_pending(datetime(2026, 8, 9, 9, 30, tzinfo=ZoneInfo("Europe/Moscow")))
    assert dispatcher.calls[0] == ("morning", "morning:2026-08-09")
    assert dispatcher.calls[1][0] == "operational"


def test_scheduler_before_morning_sends_only_operational():
    dispatcher = FakeDispatcher()
    scheduler = TelegramReportScheduler(dispatcher, timezone_name="Europe/Moscow", morning_time="09:00", operational_interval_seconds=10800)
    scheduler.run_pending(datetime(2026, 8, 9, 8, 59, tzinfo=ZoneInfo("Europe/Moscow")))
    assert [call[0] for call in dispatcher.calls] == ["operational"]
