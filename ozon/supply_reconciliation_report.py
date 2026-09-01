from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db import SessionLocal
from app.models import OzonFBOSupplyAct, OzonFBOSupplyActItem, OzonFBOSupplyDeclaredItem


DEFAULT_REPORT_PATH = Path("data/ozon/reports/ozon_fbo_supply_reconciliation_all.xlsx")


def _excel_value(value: Any) -> Any:
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(ZoneInfo("Europe/Moscow")).replace(tzinfo=None)
    return value


class OzonSupplyReconciliationExcelReport:
    """Build a complete sent-versus-accepted FBO report from the local database."""

    def __init__(self, *, session_factory: Callable[..., Any] = SessionLocal) -> None:
        self.session_factory = session_factory

    def build(self) -> tuple[bytes, int, int]:
        with self.session_factory() as session:
            declared_rows = session.query(OzonFBOSupplyDeclaredItem).order_by(
                OzonFBOSupplyDeclaredItem.supply_order_id,
                OzonFBOSupplyDeclaredItem.supply_id,
                OzonFBOSupplyDeclaredItem.sku,
            ).all()
            act_rows = session.query(OzonFBOSupplyActItem, OzonFBOSupplyAct).outerjoin(
                OzonFBOSupplyAct,
                OzonFBOSupplyAct.act_id == OzonFBOSupplyActItem.act_id,
            ).order_by(
                OzonFBOSupplyActItem.supply_order_id,
                OzonFBOSupplyActItem.supply_id,
                OzonFBOSupplyActItem.sku,
                OzonFBOSupplyActItem.act_id,
            ).all()

        details: dict[tuple[int, int, int], dict[str, Any]] = {}
        for row in declared_rows:
            key = (row.supply_order_id, row.supply_id, row.sku)
            details[key] = {
                "order_id": row.supply_order_id,
                "supply_id": row.supply_id,
                "bundle_id": row.bundle_id,
                "state": row.supply_state,
                "warehouse_id": row.storage_warehouse_id,
                "warehouse": row.storage_warehouse_name,
                "sku": row.sku,
                "product_id": row.product_id,
                "offer_id": row.offer_id,
                "name": row.name,
                "barcode": row.barcode,
                "shipment_type": row.shipment_type,
                "placement_zone": row.placement_zone,
                "pack_quantity": row.pack_quantity,
                "sent": row.declared_quantity or 0,
                "accepted": 0,
                "approved": 0,
                "defect": 0,
                "surplus": 0,
                "shortcoming": 0,
                "completed": False,
                "act_ids": set(),
                "act_numbers": set(),
                "act_types": set(),
                "act_dates": set(),
                "bundle_fetched_at": row.fetched_at,
                "act_fetched_at": None,
            }

        for item, act in act_rows:
            key = (item.supply_order_id, item.supply_id, item.sku)
            detail = details.setdefault(key, {
                "order_id": item.supply_order_id,
                "supply_id": item.supply_id,
                "bundle_id": None,
                "state": None,
                "warehouse_id": None,
                "warehouse": None,
                "sku": item.sku,
                "product_id": None,
                "offer_id": item.offer_id,
                "name": item.name,
                "barcode": item.barcode,
                "shipment_type": None,
                "placement_zone": None,
                "pack_quantity": None,
                "sent": 0,
                "accepted": 0,
                "approved": 0,
                "defect": 0,
                "surplus": 0,
                "shortcoming": 0,
                "completed": False,
                "act_ids": set(),
                "act_numbers": set(),
                "act_types": set(),
                "act_dates": set(),
                "bundle_fetched_at": None,
                "act_fetched_at": None,
            })
            detail["offer_id"] = detail["offer_id"] or item.offer_id
            detail["name"] = detail["name"] or item.name
            detail["barcode"] = detail["barcode"] or item.barcode
            act_type = str(item.act_type or "UNSPECIFIED").upper()
            if act_type == "ACCEPTANCE":
                detail["accepted"] += item.fact_quantity or 0
                detail["approved"] += item.approved_quantity or 0
            elif act_type == "DEFECT":
                detail["defect"] += item.fact_quantity or 0
            elif act_type == "SURPLUS":
                detail["surplus"] += item.fact_quantity or 0
            elif act_type == "SHORTCOMING":
                detail["shortcoming"] += item.fact_quantity or 0
            detail["act_ids"].add(item.act_id)
            detail["act_types"].add(act_type)
            detail["act_fetched_at"] = max(
                filter(None, (detail["act_fetched_at"], item.fetched_at)),
                default=None,
            )
            if act is not None:
                detail["completed"] = detail["completed"] or bool(act.is_agreement_completed)
                if act.act_number:
                    detail["act_numbers"].add(act.act_number)
                if act.act_created_date:
                    detail["act_dates"].add(act.act_created_date.isoformat())

        ordered = [details[key] for key in sorted(details)]
        detail_rows = [self._detail_row(item) for item in ordered]
        supply_rows = self._supply_rows(ordered)
        content = self._workbook(detail_rows, supply_rows)
        return content, len(detail_rows), len(supply_rows)

    def save(self, destination: str | Path = DEFAULT_REPORT_PATH) -> dict[str, Any]:
        path = Path(destination).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)
        content, product_rows, supply_rows = self.build()
        temporary = path.with_name(f".{path.name}.{uuid4().hex}.tmp")
        try:
            temporary.write_bytes(content)
            os.replace(temporary, path)
        finally:
            temporary.unlink(missing_ok=True)
        return {
            "path": str(path),
            "product_rows": product_rows,
            "supply_rows": supply_rows,
            "bytes": len(content),
        }

    @staticmethod
    def _detail_row(item: dict[str, Any]) -> tuple[Any, ...]:
        return (
            item["order_id"], item["supply_id"], item["bundle_id"], item["state"],
            item["warehouse_id"], item["warehouse"], item["sku"], item["product_id"],
            item["offer_id"], item["name"], item["barcode"], item["shipment_type"],
            item["placement_zone"], item["pack_quantity"], item["sent"], item["accepted"],
            item["approved"], item["accepted"] - item["sent"], item["defect"],
            item["surplus"], item["shortcoming"], "Да" if item["completed"] else "Нет",
            ", ".join(str(value) for value in sorted(item["act_ids"])),
            ", ".join(sorted(item["act_numbers"])),
            ", ".join(sorted(item["act_types"])),
            ", ".join(sorted(item["act_dates"])),
            item["bundle_fetched_at"], item["act_fetched_at"],
        )

    @staticmethod
    def _supply_rows(details: list[dict[str, Any]]) -> list[tuple[Any, ...]]:
        supplies: dict[tuple[int, int], dict[str, Any]] = {}
        for item in details:
            key = (item["order_id"], item["supply_id"])
            row = supplies.setdefault(key, {
                "order_id": item["order_id"], "supply_id": item["supply_id"],
                "bundle_id": item["bundle_id"], "state": item["state"],
                "warehouse_id": item["warehouse_id"], "warehouse": item["warehouse"],
                "sku_count": 0, "sent": 0, "accepted": 0, "approved": 0,
                "defect": 0, "surplus": 0, "shortcoming": 0, "completed": False,
                "act_numbers": set(), "act_dates": set(),
            })
            row["sku_count"] += 1
            for field in ("sent", "accepted", "approved", "defect", "surplus", "shortcoming"):
                row[field] += item[field]
            row["completed"] = row["completed"] or item["completed"]
            row["act_numbers"].update(item["act_numbers"])
            row["act_dates"].update(item["act_dates"])
        return [(
            row["order_id"], row["supply_id"], row["bundle_id"], row["state"],
            row["warehouse_id"], row["warehouse"], row["sku_count"], row["sent"],
            row["accepted"], row["approved"], row["accepted"] - row["sent"],
            row["defect"], row["surplus"], row["shortcoming"],
            "Да" if row["completed"] else "Нет", ", ".join(sorted(row["act_numbers"])),
            ", ".join(sorted(row["act_dates"])),
        ) for _, row in sorted(supplies.items())]

    @staticmethod
    def _workbook(detail_rows: list[tuple[Any, ...]], supply_rows: list[tuple[Any, ...]]) -> bytes:
        detail_headers = (
            "Номер заявки", "ID поставки", "Bundle ID", "Статус поставки", "ID склада Ozon",
            "Склад Ozon", "SKU", "Product ID", "Артикул продавца", "Товар", "Штрихкод",
            "Тип отгрузки", "Зона размещения", "Кол-во в упаковке", "Отправлено",
            "Принято фактически", "Согласовано Ozon", "Расхождение принято−отправлено",
            "Брак", "Излишек", "Недостача", "Приёмка завершена", "ID актов",
            "Номера актов", "Типы актов", "Даты актов", "Состав получен", "Акты получены",
        )
        supply_headers = (
            "Номер заявки", "ID поставки", "Bundle ID", "Статус поставки", "ID склада Ozon",
            "Склад Ozon", "Количество SKU", "Отправлено", "Принято фактически",
            "Согласовано Ozon", "Расхождение принято−отправлено", "Брак", "Излишек",
            "Недостача", "Приёмка завершена", "Номера актов", "Даты актов",
        )
        workbook = Workbook()
        workbook.remove(workbook.active)
        for title, headers, rows in (
            ("По товарам", detail_headers, detail_rows),
            ("По поставкам", supply_headers, supply_rows),
        ):
            sheet = workbook.create_sheet(title)
            sheet.append(headers)
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
            widths = [len(header) for header in headers]
            for row in rows:
                values = [_excel_value(value) for value in row]
                sheet.append(values)
                for index, value in enumerate(values):
                    if value is not None:
                        widths[index] = min(60, max(widths[index], len(str(value))))
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
            for index, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = min(width + 2, 60)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        return stream.getvalue()
